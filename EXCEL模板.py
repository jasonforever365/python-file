# -*- coding: utf-8 -*-
"""
Created on Fri Jan 15 07:07:53 2021

@author: zhou jian
"""


import openpyxl
wb = openpyxl.load_workbook('20200916 TF activity review template - Arab bank plc.xlsx')
#getting sheets from the workbook
print(wb.sheetnames)
#遍历表单
for sheet in wb:
    print(sheet.title)

mySheet = wb.create_sheet('mySheet')
print(wb.sheetnames)

# sheet3 = wb.get_sheet_by_name('Sheet3')
# sheet4 = wb['mySheet']
# Getting cells from the sheets
ws = wb.active
print(ws)
print(ws['A1'])
print(ws['A1'].value)

c = ws['B1']
print('Row {}, Column {} is {}'.format(c.row, c.column, c.value))
print('Cell {} is {}\n'.format(c.coordinate, c.value))

print(ws.cell(row = 1, columm = 2).value)

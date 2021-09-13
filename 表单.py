# -*- coding: utf-8 -*-
"""
Created on Tue Jan 12 21:54:47 2021

@author: zhou jian
"""


from openpyxl import Workbook
from openpyxl import load_workbook
import random
#创建excel表格
fileName = 'example.xlsx'
wb = Workbook()
#默认创建一个名为“sheet”的sheet
ws = wb.active
print('默认创建的sheet名：', ws.title)
#修改下面tab的颜色
ws.title = 'Sheet1'
#修改下面tab的颜色
ws.sheet_properties.tabColor = 'FF0000'
#创建新sheet，可以控制位置，都可以表示在这个位置之前插入
ws1 = wb.create_sheet('Mysheet1')
ws2 = wb.create_sheet('Mysheet2',0)
ws3 = wb.create_sheet('Mysheet3',-1)

#遍历所有的sheet
for sheetname in wb.sheetnames:
    print('遍历sheet：', sheetname)

#选中sheet
ws3 = wb['MySheet3'] 
print('选中sheet name:', ws3.title)
#修改active sheet
wb.active = wb['MySheet3']
ws3 = wb.active
print('当前active sheet', ws3.title)
#单元格赋值
ws3['A1']=1
#获取单元格
cell = ws3['A1']
print('A1的值', cell.value)
   
